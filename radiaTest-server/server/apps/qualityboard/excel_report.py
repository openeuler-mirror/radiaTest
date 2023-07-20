# Copyright (c) [2022] Huawei Technologies Co.,Ltd.ALL rights reserved.
# This program is licensed under Mulan PSL v2.
# You can use it according to the terms and conditions of the Mulan PSL v2.
#          http://license.coscl.org.cn/MulanPSL2
# THIS PROGRAM IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
# EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
# MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
# See the Mulan PSL v2 for more details.
####################################
# @Author  : hukun66
# @email   : hu_kun@hoperun.com
# @Date    : 2023/09/04
# @License : Mulan PSL v2
#####################################

import os
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import Reference, PieChart, BarChart3D
from openpyxl.chart.series_factory import SeriesFactory
from openpyxl.styles import PatternFill, Side, Border, Font, Alignment
from flask import current_app


# 默认字体
default_font = Font(name="宋体", size=10.5, color="000000")
# 加粗字体
bold_font = Font(name="宋体", size=10.5, color="000000", bold=True)
thin = Side(border_style='thin', color='000000')
default_border = Border(left=thin, right=thin, top=thin, bottom=thin)
default_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


def write_content(sheet, row, column, value, font=None):
    # 写入内容
    if not font:
        font = default_font
    sheet.cell(row=row, column=column).value = value
    sheet.cell(row=row, column=column).font = font
    sheet.cell(row=row, column=column).alignment = default_alignment
    sheet.cell(row=row, column=column).border = default_border


def set_border(sheet, start_row, end_row, start_column, end_column):
    # 指定区域设置边框线
    for row in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            sheet.cell(row=row, column=col).border = default_border


class BaseReport(object):
    total_col_num = 14
    report_type = "base"

    def __init__(self, product_name, round_name=None):
        self.product_name = product_name
        self.round_name = round_name
        self.save_path = Path(current_app.config.get("TMP_FILE_SAVE_PATH")).joinpath("report")
        if not self.save_path.exists():
            self.save_path.mkdir()

    def merge_cells(self, sheet, start_row, end_row=None, start_column=None, end_column=None):
        # 单元格合并
        if not end_row:
            end_row = start_row
        if not start_column:
            start_column = 1
        if not end_column:
            end_column = self.total_col_num
        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=start_column, end_column=end_column)

    @staticmethod
    def set_cell_color(sheet, row, col, color):
        # 设置单元格颜色
        sheet.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    def _main(self, wb, excel_data):
        # 生成报告逻辑必须实现
        raise Exception("子类的_main方法必须实现")

    def generate_excel_report(self, excel_data):
        name = f"{self.product_name}_{self.round_name}" if self.round_name else f"{self.product_name}"
        save_file = self.save_path.joinpath(f'{name}_{self.report_type}.xlsx')
        if save_file.exists():
            os.remove(save_file)
        # 创建一个工作簿对象
        wb = Workbook()
        try:
            self._main(wb, excel_data)
            # 报告保存
            wb.save(save_file)
            # 最后关闭文件
        finally:
            wb.close()
        return save_file


class QualityReport(BaseReport):
    total_col_num = 15
    report_type = "质量报告"

    @staticmethod
    def write_issue_list(sheet, start_row, column_header, issue_list):
        # 向sheet表写入issue列表数据
        current_row = start_row
        # 写入header信息
        for index, value in enumerate(column_header.values(), start=1):
            write_content(sheet, current_row, index, value, bold_font)
        # 写入issue信息
        for issue_info in issue_list:
            current_row = current_row + 1
            for col_index, key in enumerate(column_header.keys(), start=1):
                write_content(sheet, current_row, col_index, issue_info.get(key))
        return current_row

    def _main(self, wb, excel_data):
        ws = wb.create_sheet('总览', 0)

        row_num = 1
        self.merge_cells(ws, row_num)
        write_content(ws, row_num, 1, '版本健康度统计', font=bold_font)
        self.set_cell_color(ws, row_num, 1, '2F75B5')

        row_num += 1
        line2 = '一、问题健康度（仅关注在行版本分支），'
        branch_str = f'当前为{excel_data["branch"]}分支' if excel_data.get("branch") else '未选择分支'

        self.merge_cells(ws, row_num)
        write_content(ws, row_num, 1, line2 + branch_str + "。", font=bold_font)
        self.set_cell_color(ws, row_num, 1, 'BDD7EE')

        row_num += 1
        self.merge_cells(ws, row_num)
        status_count_str = "，".join([f"{name}{count}个" for name, count in excel_data["issue_count"]["status"].items()])
        line3 = "问题单总览：{}，共{}个issue{}".format(
            branch_str, excel_data["issue_count"]["all_count"], "，" + status_count_str if status_count_str else "")
        write_content(ws, row_num, 1, line3 + "。", font=bold_font)

        row_num += 1
        self.merge_cells(ws, row_num)
        write_content(ws, row_num, 1, '严重&block问题列表', font=bold_font)
        self.set_cell_color(ws, row_num, 1, 'FFC000')

        row_num += 1
        write_content(ws, row_num, 1, '未闭环问题总数', font=bold_font)
        self.merge_cells(ws, row_num, start_column=2, end_column=3)
        write_content(ws, row_num, 2, excel_data["issue_count"]['unclosed'], font=bold_font)

        write_content(ws, row_num, 4, '问题闭环率', font=bold_font)
        self.merge_cells(ws, row_num, start_column=5, end_column=8)
        write_content(ws, row_num, 5, excel_data["closed_rate"], font=bold_font)

        self.merge_cells(ws, row_num, start_column=9, end_column=10)
        write_content(ws, row_num, 9, '严重/阻塞问题数', font=bold_font)
        write_content(ws, row_num, 11, excel_data["issue_count"]['focus'], font=bold_font)

        write_content(ws, row_num, 12, '总DI值', font=bold_font)
        self.merge_cells(ws, row_num, start_column=13)
        write_content(ws, row_num, 13, excel_data["version_di"], font=bold_font)

        table_header = OrderedDict([("milestone", "里程碑"), ("issue_id", "任务ID"), ("issue_title", "任务标题"),
                                    ("assignee_username", "任务负责人"), ("assignee_name", "姓名"),
                                    ("assignee_enterprise", "所属企业"),
                                    ("plan_started_at", "计划开始日期"), ("deadline", "计划截止日期"),
                                    ("finished_at", "完成日期"),
                                    ("issue_type", "任务类型"), ("status", "任务状态"), ("overdue_days", "逾期情况"),
                                    ("priority_human", "优先级"), ("is_block", "是否block"), ("label", "标签")])

        # 写入block、严重issue
        row_num += 1
        row_num = self.write_issue_list(ws, row_num, table_header, excel_data["abnormal_list"]) + 2

        # 写入未闭环issue
        if excel_data["unclosed_list"]:
            self.merge_cells(ws, row_num)
            ws1 = wb.create_sheet('未闭环issue', 1)
            self.write_issue_list(ws1, 1, table_header, excel_data["unclosed_list"])
            write_content(ws, row_num, 1, '未闭环issue总览：请查看(未闭环issue sheet页面)',
                          font=Font(name="宋体", size=10.5, color="0563C1"))
            # 插入超链接sheet
            ws.cell(row=row_num, column=1).hyperlink = "#未闭环issue!A1"
            self.set_cell_color(ws, row_num, 1, 'FFFF00')
        else:
            ws1 = None

        # 统计分析部分
        row_num += 2
        self.merge_cells(ws, row_num)
        write_content(ws, row_num, 1, '二、统计分析', font=bold_font)
        self.set_cell_color(ws, row_num, 1, 'BDD7EE')

        row_num += 1
        self.merge_cells(ws, row_num)
        write_content(ws, row_num, 1, '1)相关联里程碑总览：')
        self.set_cell_color(ws, row_num, 1, '92D050')

        row_num += 1
        self.merge_cells(ws, row_num, start_column=1, end_column=3)
        write_content(ws, row_num, 1, '里程碑')

        for milestone in sorted(excel_data["milestone_count"].keys()):
            row_num += 1
            self.merge_cells(ws, row_num, start_column=1, end_column=3)
            write_content(ws, row_num, 1, milestone)

        row_num += 3
        self.merge_cells(ws, row_num)
        write_content(ws, row_num, 1, '2)特性&专项测试及对应问题单个数：')
        self.set_cell_color(ws, row_num, 1, '92D050')

        row_num += 1
        write_content(ws, row_num, 1, "特性&专项测试")
        write_content(ws, row_num, 2, "问题单总数")
        write_content(ws, row_num, 3, "未闭环问题单数量")
        write_content(ws, row_num, 4, "DI")

        sig_start_row = row_num + 1

        sig_count = dict(sorted(excel_data["sig_count"].items(), key=lambda item: item[1]["total"], reverse=True))
        for sig_name, sig_info in sig_count.items():
            row_num += 1
            write_content(ws, row_num, 1, sig_name)
            write_content(ws, row_num, 2, sig_info["total"])
            write_content(ws, row_num, 3, sig_info["unclosed"])
            write_content(ws, row_num, 4, sig_info["DI"])

        sig_end_row = row_num

        row_num += 3
        self.merge_cells(ws, row_num)
        write_content(ws, row_num, 1, '3)issue SIG 分布：')
        self.set_cell_color(ws, row_num, 1, '92D050')
        sig_chart_x_data = Reference(ws, min_row=sig_start_row, max_row=sig_end_row, min_col=1, max_col=1)
        sig_chart_y_data = Reference(ws, min_row=sig_start_row, max_row=sig_end_row, min_col=2, max_col=2)

        if sig_end_row > sig_start_row:
            # 插入饼图
            pie_chart = PieChart()
            pie_chart.width = 30
            pie_chart.height = 12
            series = SeriesFactory(sig_chart_y_data, title="SIG分布")
            pie_chart.series.append(series)
            pie_chart.set_categories(sig_chart_x_data)

            ws.add_chart(pie_chart, f"B{row_num + 3}")
            row_num += 30
        else:
            row_num += 1
            write_content(ws, row_num, 1, '暂无')
            row_num += 2

        self.merge_cells(ws, row_num)
        write_content(ws, row_num, 1, '4)每轮迭代问题单闭环情况：')
        self.set_cell_color(ws, row_num, 1, '92D050')

        row_num += 1
        write_content(ws, row_num, 1, "迭代轮次")
        write_content(ws, row_num, 2, "问题单总数")
        write_content(ws, row_num, 3, "完成验收问题单个数")
        write_content(ws, row_num, 4, "闭环率")

        for milestone in sorted(excel_data["milestone_count"].keys()):
            row_num += 1
            write_content(ws, row_num, 1, milestone)
            write_content(ws, row_num, 2, excel_data["milestone_count"][milestone]["total"])
            write_content(ws, row_num, 3, excel_data["milestone_count"][milestone]["closed"])
            write_content(ws, row_num, 4, excel_data["milestone_count"][milestone]["closed_rate"])

        # 生成柱状图
        row_num += 3
        self.merge_cells(ws, row_num)
        write_content(ws, row_num, 1, '5)issue统计(柱状图)：')
        self.set_cell_color(ws, row_num, 1, '92D050')
        # 责任人分布
        row_num += 1
        self.merge_cells(ws, row_num)
        write_content(ws, row_num, 1, '责任人分布：', font=bold_font)
        row_num += 1
        assignee_count = excel_data["issue_count"]["assignee_count"]

        write_content(ws, row_num, 1, "责任人")
        write_content(ws, row_num, 2, "数量")
        assignee_start_row = row_num + 1
        for assignee, count in assignee_count.items():
            row_num += 1
            write_content(ws, row_num, 1, assignee)
            write_content(ws, row_num, 2, count)
        assignee_end_row = row_num

        assignee_x_data = Reference(ws, min_row=assignee_start_row, max_row=assignee_end_row, min_col=1, max_col=1)
        assignee_y_data = Reference(ws, min_row=assignee_start_row, max_row=assignee_end_row, min_col=2, max_col=2)
        assignee_count_chart = BarChart3D()
        assignee_count_chart.width = 35
        assignee_count_chart.height = 20
        assignee_count_chart.title = "责任人分布"
        assignee_count_chart.type = "bar"
        series = SeriesFactory(assignee_y_data)
        assignee_count_chart.series.append(series)
        assignee_count_chart.set_categories(assignee_x_data)
        ws.add_chart(assignee_count_chart, f"C{assignee_start_row - 1}")

        # 柱状图固定约43个单元格的高度，提前预留
        if assignee_end_row - assignee_start_row + 1 < 43:
            row_num = assignee_start_row + 43
        # SIG组分布
        sig_count_chart = BarChart3D()
        sig_count_chart.width = 35
        sig_count_chart.height = 20
        sig_count_chart.title = "SIG分布"
        sig_count_chart.type = "bar"
        series = SeriesFactory(sig_chart_y_data)
        sig_count_chart.series.append(series)
        sig_count_chart.set_categories(sig_chart_x_data)
        ws.add_chart(sig_count_chart, f"C{row_num + 1}")
        # sig组无需填入数据只需汇总柱状图，故高度固定43
        row_num += 43 + 2
        # 优先级分布
        self.merge_cells(ws, row_num)
        write_content(ws, row_num, 1, '优先级分布：', font=bold_font)
        row_num += 1
        priority_count = excel_data["issue_count"]["priority"]
        write_content(ws, row_num, 1, "优先级")
        write_content(ws, row_num, 2, "数量")
        priority_start_row = row_num + 1
        for priority, count in priority_count.items():
            row_num += 1
            write_content(ws, row_num, 1, priority)
            write_content(ws, row_num, 2, count)
        priority_end_row = row_num
        priority_x_data = Reference(ws, min_row=priority_start_row, max_row=priority_end_row, min_col=1, max_col=1)
        priority_y_data = Reference(ws, min_row=priority_start_row, max_row=priority_end_row, min_col=2, max_col=2)
        priority_count_chart = BarChart3D()
        priority_count_chart.width = 35
        priority_count_chart.height = 20
        priority_count_chart.title = "优先级分布"
        priority_count_chart.type = "bar"
        series = SeriesFactory(priority_y_data)
        priority_count_chart.series.append(series)
        priority_count_chart.set_categories(priority_x_data)
        ws.add_chart(priority_count_chart, f"C{priority_start_row - 1}")
        # 优先级数据仅四行，故也固定为柱状图高度
        row_num += 40
        # 任务状态分布
        self.merge_cells(ws, row_num)
        write_content(ws, row_num, 1, '任务状态分布：', font=bold_font)
        row_num += 1
        status_count = excel_data["issue_count"]["status"]
        write_content(ws, row_num, 1, "任务状态")
        write_content(ws, row_num, 2, "数量")
        status_start_row = row_num + 1
        for status, count in status_count.items():
            row_num += 1
            write_content(ws, row_num, 1, status)
            write_content(ws, row_num, 2, count)
        status_end_row = row_num
        status_x_data = Reference(ws, min_row=status_start_row, max_row=status_end_row, min_col=1, max_col=1)
        status_y_data = Reference(ws, min_row=status_start_row, max_row=status_end_row, min_col=2, max_col=2)
        status_count_chart = BarChart3D()
        status_count_chart.width = 35
        status_count_chart.height = 20
        status_count_chart.title = "任务状态分布"
        status_count_chart.type = "bar"
        series = SeriesFactory(status_y_data)
        status_count_chart.series.append(series)
        status_count_chart.set_categories(status_x_data)
        ws.add_chart(status_count_chart, f"C{status_start_row - 1}")
        # 任务状态数据也小于42，故也固定为柱状图高度
        row_num += 43

        # 全部设置边框
        set_border(ws, 1, row_num, 1, self.total_col_num)
        # 设置列宽
        ws.column_dimensions.width = 15
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12

        ws.column_dimensions['O'].width = 15

        ws.column_dimensions['L'].width = 10
        if ws1:
            ws1.column_dimensions.width = 15
            ws1.column_dimensions['A'].width = 32
            ws1.column_dimensions['B'].width = 10
            ws1.column_dimensions['C'].width = 40
            ws1.column_dimensions['D'].width = 15
            ws1.column_dimensions['E'].width = 15

            ws1.column_dimensions['G'].width = 12
            ws1.column_dimensions['H'].width = 12
            ws1.column_dimensions['I'].width = 12
            ws1.column_dimensions['L'].width = 10
            ws1.column_dimensions['O'].width = 15


class ATReport(BaseReport):
    total_col_num = 8
    report_type = "AT报告"
    overview_table_header = OrderedDict(
            [("build", "构建记录"), ("total", "总数"), ("success", "成功"), ("running", "正在执行"),
             ("failure", "失败"), ("block", "阻塞"), ("test_duration", "测试时长"), ("detail", "详细数据")])

    detail_table_header = OrderedDict(
            [("test", "测试项"), ("aarch64_res_status", "aarch64 结果"), ("aarch64_res_log", "aarch64 结果日志"),
             ("aarch64_start_time", "aarch64 开始时间"), ("aarch64_end_name", "aarch64 结束时间"),
             ("aarch64_test_duration", "aarch64 测试时长"), ("aarch64_failedmodule_name", "aarch64 失败模块名"),
             ("aarch64_failedmodule_log", "aarch64 失败模块日志"),
             ("x86_64_res_status", "x86_64 结果"), ("x86_64_res_log", "x86_64 结果日志"),
             ("x86_64_start_time", "x86_64 开始时间"), ("x86_64_end_name", "x86_64 结束时间"),
             ("x86_64_test_duration", "x86_64 测试时长"), ("x86_64_failedmodule_name", "x86_64 失败模块名"),
             ("x86_64_failedmodule_log", "x86_64 失败模块日志"),
             ])

    def write_at_detail_list(self, workbook, sheet_name, detail_list, sheet_index):
        current_sheet_index = sheet_index + 1
        detail_sheet = workbook.create_sheet(sheet_name, current_sheet_index)
        row_index = 1
        # 写入header信息
        for index, value in enumerate(self.detail_table_header.values(), start=1):
            write_content(detail_sheet, row_index, index, value, bold_font)

        # 写入详情列表
        for detail in detail_list:
            row_index += 1
            for col_index, key in enumerate(self.detail_table_header.keys(), start=1):
                write_content(detail_sheet, row_index, col_index, detail.get(key))

        # 设置列宽
        detail_sheet.column_dimensions['A'].width = 32
        # arm
        detail_sheet.column_dimensions['B'].width = 15
        detail_sheet.column_dimensions['C'].width = 32
        detail_sheet.column_dimensions['D'].width = 18
        detail_sheet.column_dimensions['E'].width = 18
        detail_sheet.column_dimensions['F'].width = 8
        detail_sheet.column_dimensions['G'].width = 18
        detail_sheet.column_dimensions['H'].width = 32
        # x86
        detail_sheet.column_dimensions['I'].width = 15
        detail_sheet.column_dimensions['J'].width = 32
        detail_sheet.column_dimensions['K'].width = 18
        detail_sheet.column_dimensions['L'].width = 18
        detail_sheet.column_dimensions['M'].width = 8
        detail_sheet.column_dimensions['N'].width = 18
        detail_sheet.column_dimensions['O'].width = 32

        return current_sheet_index

    def _main(self, wb, res_list):
        # 向sheet表写入at列表数据
        sheet_index = 0
        ws = wb.create_sheet('AT历史记录总览', sheet_index)
        row_index = 1

        detail_col = 8
        # 写入header信息
        for index, value in enumerate(self.overview_table_header.values(), start=1):
            write_content(ws, row_index, index, value, bold_font)
        # 写入build信息
        for res_info in res_list:
            row_index = row_index + 1
            for col_index, key in enumerate(self.overview_table_header.keys(), start=1):
                write_content(ws, row_index, col_index, res_info.get(key))
            # 写入详情数据
            detail_list = res_info.get("detail_list")
            if detail_list:
                # 写入详情跳转链接
                write_content(ws, row_index, detail_col, "请点击", font=Font(name="宋体", size=10.5, color="0563C1"))
                # 插入超链接sheet
                build_name = res_info.get("build")
                ws.cell(row=row_index, column=detail_col).hyperlink = f"#'{build_name}'!A1"
                sheet_index = self.write_at_detail_list(wb, build_name, detail_list, sheet_index)
            else:
                write_content(ws, row_index, detail_col, "暂无")
        # 全部设置边框
        set_border(ws, 1, row_index, 1, self.total_col_num)
        # 设置列宽
        ws.column_dimensions['A'].width = 25

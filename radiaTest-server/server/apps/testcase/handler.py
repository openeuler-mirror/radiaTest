# Copyright (c) [2022] Huawei Technologies Co.,Ltd.ALL rights reserved.
# This program is licensed under Mulan PSL v2.
# You can use it according to the terms and conditions of the Mulan PSL v2.
# http://license.coscl.org.cn/MulanPSL2
# THIS PROGRAM IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
# EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
# MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
# See the Mulan PSL v2 for more details.
####################################
# Author : MDS_ZHR
# email : 331884949@qq.com
# Date : 2022/12/13 14:00:00
# License : Mulan PSL v2
#####################################
# 用例管理(Testcase)相关接口的handler层

import abc
import math
import os
import datetime
import time
from typing import List
import openpyxl
from openpyxl.styles import Alignment

from flask import jsonify, g, current_app, request
import sqlalchemy
from sqlalchemy.exc import IntegrityError, SQLAlchemyError


from server import db, redis_client
from server.utils.redis_util import RedisKey
from server.utils.response_util import RET
from server.utils.db import Insert, collect_sql_error, Edit
from server.utils.page_util import PageUtil
from server.model.testcase import (
    CaseNode,
    Suite,
    Case,
    SuiteDocument,
    Baseline,
    CaseResult
)
from server.model.framework import GitRepo
from server.model.task import Task, TaskMilestone, TaskManualCase
from server.model.celerytask import CeleryTask
from server.model.organization import Organization
from server.model.milestone import Milestone
from server.schema.testcase import (
    CaseNodeBaseSchema, 
    AddCaseCommitSchema,
    CaseNodeBodySchema
)
from server.schema.celerytask import CeleryTaskUserInfoSchema
from server.utils.file_util import MarkdownImportFile, ZipImportFile, ExcelImportFile
from server.utils.sheet import Excel
from server.utils.permission_utils import GetAllByPermission
from server.utils.md_util import MdUtil
from celeryservice.tasks import resolve_testcase_file, resolve_testcase_set
from celeryservice.sub_tasks import create_case_node_multi_select


class CaseImportHandler:
    CONVERT_METHOD = {
        "md": MdUtil.df2md,
        "xlsx": MdUtil.md2wb,
    }

    def __init__(self, file):
        try:
            try:
                self.case_file = ExcelImportFile(file)
            except TypeError:
                self.case_file = MarkdownImportFile(file)
            
            if self.case_file.filetype:
                testcase_file_path = os.path.join(current_app.config.get("TMP_FILE_SAVE_PATH"), "testcase")
                self.case_file.file_save(testcase_file_path)
            else:
                mesg = "Filetype of {}.{} is not supported".format(
                        self.case_file.filename,
                        self.case_file.filetype,
                    )
                raise RuntimeError(mesg)
        
        except RuntimeError as e:
            current_app.logger.error(str(e))
            if os.path.exists(self.case_file.filepath):
                self.case_file.file_remove()
            raise e

    def import_case(self, group_id=None, case_node_id=None):
        permission_type = "org"
        if group_id:
            try:
                _ = int(group_id)
                permission_type = "group"
            except (ValueError, TypeError):
                group_id = None
        else:
            group_id = None

        _task = resolve_testcase_file.delay(
            self.case_file.filepath,
            CeleryTaskUserInfoSchema(
                auth=request.headers.get("authorization"),
                user_id=g.user_id,
                group_id=group_id,
                org_id=redis_client.hget(
                    RedisKey.user(g.user_id),
                    'current_org_id'
                ),
                permission_type=permission_type,
            ).__dict__,
            case_node_id
        )

        if not _task:
            return jsonify(
                error_code=RET.SERVER_ERR, 
                error_msg="could not send task to resolve file"
            )

        _ = Insert(
            CeleryTask,
            {
                "tid": _task.task_id,
                "status": "PENDING",
                "object_type": "testcase_resolve",
                "description": f"resolve testcase {self.case_file.filepath}",
                "user_id": g.user_id
            }
        ).single(CeleryTask, '/celerytask')

        return jsonify(error_code=RET.OK, error_msg="OK")

    def convert(self, to=None):
        """convert md => xlsx or xlsx => md

        Args:
            to (str): the target filetype to transfer
        
        Return:
            filepath (str): the convert file save path 

        """
        testcase_file_path = os.path.join(current_app.config.get("TMP_FILE_SAVE_PATH"), "export")
        os.makedirs(testcase_file_path, exist_ok=True)
        try:                       
            if to == "md":
                kwargs = {
                    "df": Excel(self.case_file.filetype).load(self.case_file.filepath),
                    "md_path": "{}{}.md".format(
                        testcase_file_path,
                        self.case_file.filename,
                    )
                }
            elif to == "xlsx":
                with open(self.case_file.filepath, "r") as f:
                    _wb_path = "{}{}.xlsx".format(
                        testcase_file_path,
                        self.case_file.filename,
                    )

                    kwargs = {
                        "md_content": f.read(), 
                        "wb_path": _wb_path,
                        "sheet_name": self.case_file.filename,
                    }
                
                wb = openpyxl.Workbook()
                wb.create_sheet(self.case_file.filename)
                wb.save(_wb_path)
            
            return self.CONVERT_METHOD[to](**kwargs)

        except KeyError as e:
            current_app.logger.error(str(e))
            raise RuntimeError(f"convert to {to} is not supported") from e


class CaseNodeHandler:
    @staticmethod
    @collect_sql_error
    def get(case_node_id, query):
        case_node = CaseNode.query.filter_by(id=case_node_id).first()
        if not case_node:
            return jsonify(
                error_code=RET.NO_DATA_ERR, 
                error_msg="case_node does not exists"
            )

        filter_params = GetAllByPermission(CaseNode, org_id=query.org_id).get_filter()
        return_data = CaseNodeBaseSchema(**case_node.__dict__).dict()
        filter_params.append(CaseNode.parent.contains(case_node))
        if query.title:
            filter_params.append(CaseNode.title.like(f'%{query.title}%'))
        
        children = CaseNode.query.filter(*filter_params).all()

        return_data["children"] = [child.to_json() for child in children]

        source = list()
        cur = case_node
        while cur:
            if not cur.parent.all():
                source.append(cur.id)
                break
            if len(cur.parent.all()) > 1:
                raise RuntimeError(
                    "case_node should not have parents beyond one")

            source.append(cur.id)
            cur = cur.parent[0]

        return_data["source"] = source
        if case_node.type == 'case' and source:
            case_result = CaseNodeHandler.get_case_result(case_node.case_id, source[1])
            return_data["result"] = case_result
        return jsonify(error_code=RET.OK, error_msg="OK", data=return_data)

    @staticmethod
    @collect_sql_error
    def get_case_set_node(query):
        org_id = redis_client.hget(RedisKey.user(g.user_id), "current_org_id")
        filter_param = [
            CaseNode.type == "suite",
            CaseNode.in_set.is_(True),
            CaseNode.org_id == org_id
        ]
        if query.title:
            filter_param.append(
                CaseNode.title == query.title,
            )
        elif query.suite_id:
            filter_param.append(
                CaseNode.suite_id == query.suite_id,
            )
            
        case_node = CaseNode.query.filter(*filter_param).first()
        if not case_node:
            return jsonify(
                error_code=RET.NO_DATA_ERR, 
                error_msg="case node does not exist."
            )
        return_data = CaseNodeBaseSchema(**case_node.__dict__).dict()
        children = CaseNode.query.filter(
            CaseNode.in_set.is_(True),
            CaseNode.org_id == org_id,
            CaseNode.parent.contains(case_node),
        ).all()
        return_data["children"] = [child.to_json() for child in children]

        return jsonify(
            error_code=RET.OK,
            error_msg="OK",
            data=return_data
        )

    @staticmethod
    @collect_sql_error
    def get_case_result(case_id, root_case_node_id):
        case_result = None
        case = Case.query.get(case_id)
        task = Task.query.filter_by(case_node_id=root_case_node_id, is_delete=False).order_by(
            Task.create_time.desc()).first()
        if task and not task.is_manage_task:
            task_milestone = TaskMilestone.query.filter(
                TaskMilestone.task_id == task.id,
                TaskMilestone.cases.contains(case)
            ).first()
            if case.usabled:
                case_result = task_milestone.job_result
                if task_milestone.job_result == 'block':
                    case_result = 'failed'
                elif task_milestone.job_result == 'done':
                    case_result = 'success'
            else:
                case_result = 'running'
                manual_case = TaskManualCase.query.filter_by(task_milestone_id=task_milestone.id,
                                                             case_id=case.id).first()
                if manual_case:
                    case_result = manual_case.case_result
        return case_result

    @staticmethod
    @collect_sql_error
    def get_roots(query, workspace=None):
        filter_params = GetAllByPermission(CaseNode, workspace, org_id=query.org_id).get_filter()
        filter_params.append(CaseNode.is_root.is_(True))
        for key, value in query.dict().items():
            if not value:
                continue
            if key == 'title':
                filter_params.append(CaseNode.title.like(f'%{value}%'))
            if key == 'group_id':
                filter_params.append(CaseNode.group_id == value)
                filter_params.append(CaseNode.permission_type == 'group')
            if key == 'org_id':
                filter_params.append(CaseNode.org_id == value)
                filter_params.append(CaseNode.permission_type == 'org')
        case_nodes = CaseNode.query.filter(*filter_params).all()
        return_data = [case_node.to_json() for case_node in case_nodes]
        return jsonify(error_code=RET.OK, error_msg="OK", data=return_data)

    @staticmethod
    @collect_sql_error
    def create(body):
        _body = body.__dict__
        _body.update({"creator_id": g.user_id})
        _body.update({"org_id": redis_client.hget(RedisKey.user(g.user_id), 'current_org_id')})
        if not body.parent_id:
            case_node_id = Insert(CaseNode, body.__dict__).insert_id()
            return jsonify(error_code=RET.OK, error_msg="OK", data=case_node_id)

        parent = CaseNode.query.filter_by(id=body.parent_id).first()
        if not parent:
            return jsonify(error_code=RET.NO_DATA_ERR, error_msg="parent node does not exist")
        
        _body.update({
            "permission_type": parent.permission_type,
            "in_set": parent.in_set,
            "group_id": parent.group_id,
        })

        root_case_node = CaseNodeHandler.get_root_case_node(body.parent_id)
        task = Task.query.filter(
            Task.case_node_id == root_case_node.id,
            Task.accomplish_time.is_(None),
            Task.is_delete.is_(False),
            Task.is_manage_task.is_(False),
            ).first()
        if task:
            return jsonify(
                error_code=RET.DATA_EXIST_ERR,
                error_msg="task exists,not allowed to create a new directory/suite"
            )

        if root_case_node.type == "baseline":
            _body.update({"baseline_id": root_case_node.baseline_id})
            if _body.get("type") == "case":
                _body["case_result"] = "pending"

        if _body.get("multiselect"):
            create_case_node_multi_select.delay(_body)
        else:
            for child in parent.children:
                if _body["title"] == child.title:
                    return jsonify(
                        error_code=RET.OK,
                        error_msg="Title {} is already exist".format(
                            _body["title"]
                        ),
                        data=child.id
                    )

            case_node = CaseNode.query.filter_by(
                id=Insert(
                    CaseNode,
                    _body,
                ).insert_id()
            ).first()

            case_node.parent.append(parent)
            case_node.add_update()

        return jsonify(error_code=RET.OK, error_msg="OK")

    @staticmethod
    @collect_sql_error
    def update(case_node_id, body):
        case_node = CaseNode.query.filter_by(id=case_node_id).first()

        current_org_id = int(
            redis_client.hget(RedisKey.user(g.user_id), 'current_org_id')
        )

        if current_org_id != case_node.org_id:
            return jsonify(error_code=RET.VERIFY_ERR, error_msg="No right to edit")

        if not case_node:
            return jsonify(error_code=RET.NO_DATA_ERR, error_msg="case_node does not exist")

        case_node.title = body.title

        case_node.add_update()

        return jsonify(error_code=RET.OK, error_msg="OK")

    @staticmethod
    @collect_sql_error
    def delete(case_node_id):
        case_node = CaseNode.query.filter_by(id=case_node_id).first()
        if not case_node:
            return jsonify(error_code=RET.NO_DATA_ERR, error_msg="case_node does not exist")

        current_org_id = int(
            redis_client.hget(RedisKey.user(g.user_id), 'current_org_id')
        )

        if current_org_id != case_node.org_id:
            return jsonify(error_code=RET.VERIFY_ERR, error_msg="No right to delete")

        # 解决级联删除未生效问题
        _baseline_id = case_node.baseline_id

        case_id_list = []
        if case_node.in_set:
            if case_node.type == "case" and case_node.case_id:
                case_id_list.append(case_node.case_id)
            else:
                for child in case_node.children:
                    if child.case_id:
                        case_id_list.append(child.case_id)
        db.session.delete(case_node)
        if case_node.type == "baseline" and _baseline_id:
            baseline = Baseline.query.filter_by(id=_baseline_id).first()
            db.session.delete(baseline)
        # 级联删除所有未关联用例
        for case_id in case_id_list:
            if CaseNode.query.filter_by(in_set=False, case_id=case_id).count() == 0:
                Case.query.filter_by(id=case_id).delete()
        try:
            db.session.commit()
        except (IntegrityError, SQLAlchemyError) as e:
            db.session.rollback()
            raise e

        return jsonify(error_code=RET.OK, error_msg="OK")

    @staticmethod
    @collect_sql_error
    def import_case_set(file, group_id):
        zip_case_set = ZipImportFile(file)
        uncompressed_filepath = None
        try:
            if zip_case_set.filetype:
                testcase_file_path = os.path.join(current_app.config.get("TMP_FILE_SAVE_PATH"), "testcase")
                zip_case_set.file_save(testcase_file_path)
                uncompressed_filepath = "{}/{}".format(
                    os.path.dirname(zip_case_set.filepath),
                    f"{zip_case_set.filename}_" + datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                )
                uncompress_flag = zip_case_set.uncompress(uncompressed_filepath)
                # 清理压缩包
                zip_case_set.clean_and_delete()
                if uncompress_flag is False:
                    return jsonify(error_code=RET.BAD_REQ_ERR, error_msg="解压失败，请检查压缩文件是否符合规范！")

                permission_type = "org"
                if group_id:
                    try:
                        _ = int(group_id)
                        permission_type = "group"
                    except (ValueError, TypeError):
                        group_id = None
                else:
                    group_id = None
                _task = resolve_testcase_set.delay(
                    zip_case_set.filepath,
                    uncompressed_filepath,
                    CeleryTaskUserInfoSchema(
                        auth=request.headers.get("authorization"),
                        user_id=g.user_id,
                        group_id=group_id,
                        org_id=redis_client.hget(
                            RedisKey.user(g.user_id),
                            'current_org_id'
                        ),
                        permission_type=permission_type,
                    ).__dict__,
                )

                _ = Insert(
                    CeleryTask,
                    {
                        "tid": _task.task_id,
                        "status": "PENDING",
                        "object_type": "caseset_resolve",
                        "description": "Import a set of testcases",
                        "user_id": g.user_id
                    }
                ).single(CeleryTask, '/celerytask')

                return jsonify(error_code=RET.OK, error_msg="OK")

            else:
                # 清理压缩包
                zip_case_set.clean_and_delete()
                return jsonify(error_code=RET.FILE_ERR, error_msg="filetype is not supported")

        except RuntimeError as e:
            current_app.logger.error(str(e))
            # 清理压缩包和解压目录
            zip_case_set.clean_and_delete(uncompressed_filepath)

            return jsonify(error_code=RET.SERVER_ERR, error_msg=str(e))

    @staticmethod
    @collect_sql_error
    def get_all_suite(case_node_id):
        _case_node = CaseNode.query.get(case_node_id)
        res_ids = []
        res_items = []

        def get_children(case_node: CaseNode):
            if case_node.type != 'suite' and case_node.type != "case":
                children = CaseNode.query.filter(CaseNode.parent.contains(case_node)).all()
                for child in children:
                    get_children(child)
            elif case_node.type == "suite":
                res_ids.append(case_node.suite_id)
                res_items.append(Suite.query.get(case_node.suite_id))

        if _case_node:
            get_children(_case_node)
        return res_ids, res_items

    @staticmethod
    @collect_sql_error
    def get_all_case(case_node_id):
        _case_node = CaseNode.query.get(case_node_id)
        res_ids = []
        res_items = []

        def get_children(case_node: CaseNode):
            if case_node.type != 'case':
                children = CaseNode.query.filter(CaseNode.parent.contains(case_node)).all()
                for child in children:
                    get_children(child)
            else:
                res_ids.append(case_node.case_id)
                res_items.append(Case.query.get(case_node.case_id))

        if _case_node:
            get_children(_case_node)
        return res_ids, res_items

    @staticmethod
    @collect_sql_error
    def get_root_case_node(case_node_id):
        _case_node = CaseNode.query.get(case_node_id)
        root_case_node = []

        def get_parent(case_node: CaseNode):
            if not case_node.is_root:
                parent = CaseNode.query.filter(CaseNode.children.contains(case_node)).first()
                get_parent(parent)
            else:
                root_case_node.append(case_node)

        if _case_node:
            get_parent(_case_node)
        return root_case_node[0]


    @staticmethod
    @collect_sql_error
    def get_caseset_filter(_type, _id, current_org_id):
        filter_group_params = list()
        filter_org_params = list()
        filter_params = [
            CaseNode.title == "用例集",
            CaseNode.is_root.is_(True),
            CaseNode.type == "directory",
        ]

        if _type == "org":
            filter_params.append(CaseNode.org_id == _id)
            filter_params.append(CaseNode.permission_type == _type)
        else:
            filter_group_params.append(CaseNode.group_id == _id)
            filter_group_params.append(CaseNode.permission_type == _type)

            filter_org_params.append(CaseNode.org_id == current_org_id)
            filter_org_params.append(CaseNode.permission_type == "org")
        return filter_params, filter_group_params, filter_org_params


    @staticmethod
    @collect_sql_error
    def get_caseset_return_data(current_org_id, filter_params, filter_group_params,\
            filter_org_params, return_data):

        current_org_name = Organization.query.filter_by(
            id = current_org_id
        ).first().name
        
        if filter_org_params or filter_group_params:
            filter_org_params.extend(filter_params)
            _caseset_org = CaseNode.query.filter(*filter_org_params).first()
            if _caseset_org:
                return_data.update({
                    "org": {
                        "name": current_org_name,
                        "id": current_org_id,
                        "children": [
                            child.to_json() for child in _caseset_org.children
                        ]
                    }
                })
            filter_group_params.extend(filter_params)
            _caseset_group = CaseNode.query.filter(*filter_group_params).first()
            if _caseset_group:
                return_data["group"]["children"] = [
                    child.to_json() for child in _caseset_group.children
                ]
        else:
            _caseset = CaseNode.query.filter(*filter_params).first()
            if _caseset:   
                return_data["org"]["children"] = [
                    child.to_json() for child in _caseset.children
                ]
        return return_data


    @staticmethod
    @collect_sql_error
    def get_caseset_children(_type, _table, _id):

        if _type not in ["org", "group"]:
            return jsonify(
                error_code=RET.NO_DATA_ERR,
                error_msg=f"{_type} is invalid.",
            )
        
        _item = _table.query.filter_by(id=_id).first()
        if not _item:
            return jsonify(
                error_code=RET.NO_DATA_ERR,
                error_msg=f"{_type} does not exist",
            )
        return_data = {
            _type: {
                "name": _item.name,
                "id": _id,
                "children": []
            }
        }
        
        filters = CaseNodeHandler.get_caseset_filter(
            _type, 
            _id,
            _id
        )
        
        return_data = CaseNodeHandler.get_caseset_return_data(
            _id,
            filters[0],
            filters[1], 
            filters[2], 
            return_data
        )

        return jsonify(
            error_code=RET.OK,
            error_msg="OK",
            data=return_data,
        )


    @staticmethod
    @collect_sql_error
    def get_task(case_node_id, res_type=None):
        from server.apps.task.handlers import HandlerTask
        task = Task.query.filter_by(
            case_node_id=case_node_id,
            is_delete=False
        ).order_by(Task.create_time.desc()).first()
        if task:         
            return HandlerTask.get(task.id, res_type=res_type)
        else:
            return jsonify(error_code=RET.OK, error_msg="OK", data=[])


    @staticmethod
    @collect_sql_error
    def create_relate_case_node(nodes, case_node, parent_id, 
            suite_case_node=None, case_case_node=None):
        flag = False
        if suite_case_node:
            check_suite_case_node = CaseNode.query.filter(
                CaseNode.suite_id == suite_case_node.suite_id,
                CaseNode.type == "suite",
                CaseNode.baseline_id == case_node.baseline_id,
            ).first()      
            if not check_suite_case_node:
                flag = True
                node_body = suite_case_node.to_json() 
            else:
                _id = check_suite_case_node.id
        if case_case_node:
            check_case_case_node = CaseNode.query.filter(
                CaseNode.case_id == case_case_node.case_id,
                CaseNode.type == "case",
                CaseNode.baseline_id == case_node.baseline_id,
            ).first()      
            if not check_case_case_node:
                flag = True
                node_body = case_case_node.to_json()
            else:
                _id = check_case_case_node.id

        if flag:
            if node_body["is_root"] is True:
                    node_body.update({
                        "is_root": False
                    })                         
            node_body.pop("id")
            node_body.update({
                "baseline_id":case_node.baseline_id,
                "creator_id": g.user_id,
                "case_node_id": suite_case_node.id if suite_case_node else case_case_node.id,
                "title": suite_case_node.title if suite_case_node else case_case_node.title,
                "milestone_id": case_node.milestone_id,
                "permission_type": case_node.permission_type,
            })
            if case_node.permission_type == "group":
                node_body.update({
                    "group_id": case_node.group_id
            })
            _id = Insert(CaseNode, node_body).insert_id(CaseNode, '/case_node')

            parent = CaseNode.query.filter_by(id=parent_id).first()
            child = CaseNode.query.filter_by(id=_id).first()
            child.parent.append(parent)
            child.add_update() 
        
        nodes.append(_id)
        return _id, nodes

    @staticmethod
    @collect_sql_error
    def relate(case_node_id, body):
        nodes = list()
        case_node = CaseNode.query.filter_by(
            id=case_node_id).first()
        if not case_node:
            return jsonify(error_code=RET.NO_DATA_ERR, error_msg="Case-node is not exist.")
        
        suite_case_node = CaseNode.query.filter(
            CaseNode.suite_id == body.suite_id,
            CaseNode.type == "suite",
            CaseNode.baseline_id.is_(None),
            ).first()
        if not suite_case_node:
            return jsonify(error_code=RET.NO_DATA_ERR, error_msg="suite_case_node is not exist.")

        resp = CaseNodeHandler.create_relate_case_node(
            nodes, case_node, case_node_id, suite_case_node=suite_case_node
        )
        if type(resp) == tuple:
            suite_id, nodes = resp
        else:
            return resp

        for case_id in body.case_ids:
            case_case_node = CaseNode.query.filter(
                CaseNode.case_id == case_id,
                CaseNode.type == "case",
                CaseNode.baseline_id.is_(None),
                ).first()
            if not case_case_node:
                return jsonify(error_code=RET.NO_DATA_ERR, error_msg="case_case_node is not exist.")

            resp = CaseNodeHandler.create_relate_case_node(
                nodes, case_node, suite_id, case_case_node=case_case_node
            )
            if type(resp) == tuple:
                case_id, nodes = resp
            else:
                return resp
        return jsonify(error_code=RET.OK, error_msg="OK", data=nodes)

    @staticmethod
    @collect_sql_error
    def create_task_with_casenode(case_node, body):
        if case_node.type == 'case':
            body.automatic = True

        from server.apps.task.handlers import HandlerTask
        task = Task.query.filter_by(
            case_node_id=case_node.id,
            is_delete=False
        ).order_by(Task.create_time.desc()).first()
        if not task:
            return HandlerTask.create(body)
        else:
            return jsonify(error_code=RET.VERIFY_ERR, error_msg="task of case-node is exist.")

    @staticmethod
    def get_timestamp(date):
        return time.mktime(date.timetuple())

    @staticmethod
    @collect_sql_error
    def get_suite_casenode(baseline_id, case_node_id):
        case_node = CaseNode.query.filter_by(id=case_node_id).first()
        if not case_node:
            return jsonify(error_code=RET.VERIFY_ERR, error_msg="case-node is not exist.")
        
        filter_params = [
            CaseNode.baseline_id == baseline_id,
            CaseNode.type == "suite",
        ]
        suite_nodes = CaseNode.query.join(Suite).filter(*filter_params).all()

        return suite_nodes
 

    @staticmethod
    @collect_sql_error
    def get_case_node(milestone_id):
        case_nodes = CaseNode.query.filter_by(milestone=milestone_id).all()
        return_data = [CaseNodeHandler.get_case_node_process(case_node) for case_node in case_nodes]
        return jsonify(error_code=RET.OK, error_msg="OK", data=return_data)

    @staticmethod
    @collect_sql_error
    def get_product_case_node(product_id):
        _filter = [
            Milestone.type == 'update',
            Milestone.product_id == product_id,
            CaseNode.milestone == Milestone.id
        ]
        case_nodes = CaseNode.query.filter(*_filter).all()
        return_data = [CaseNodeHandler.get_case_node_process(case_node) for case_node in case_nodes]
        return jsonify(error_code=RET.OK, error_msg="OK", data=return_data)

    @staticmethod
    @collect_sql_error
    def get_case_node_process(case_node):
        case_ids, cases = CaseNodeHandler.get_all_case(case_node.id)
        root_case_node = CaseNodeHandler.get_root_case_node(case_node.id)
        success_count = 0
        for case_id in case_ids:
            if CaseNodeHandler.get_case_result(case_id, root_case_node.id) == 'success':
                success_count = success_count + 1
        progress = '0%' if len(cases) == 0 else str(int(float(success_count) / len(cases) * 100)) + '%'
        case_node = case_node.to_json()
        case_node['progress'] = progress
        return case_node


class SuiteHandler:
    @staticmethod
    @collect_sql_error
    def create(body):
        _body = body.__dict__
        _body.update({"creator_id": g.user_id})
        _body.update({"org_id": redis_client.hget(RedisKey.user(g.user_id), 'current_org_id')})
        _id = Insert(Suite, _body).insert_id(Suite, "/suite")
        return jsonify(error_code=RET.OK, error_msg="OK", data={"id": _id})


class CaseHandler:
    @staticmethod
    @collect_sql_error
    def create(body):
        _body = body.__dict__

        _suite = Suite.query.filter_by(name=_body.get("suite")).first()
        if not _suite:
            return jsonify(
                error_code=RET.PARMA_ERR,
                error_msg="The suite {} is not exist".format(
                    _body.get("suite")
                )
            )
        _body["suite_id"] = _suite.id
        _body.pop("suite")
        _body.update({"creator_id": g.user_id})
        _body.update({"org_id": redis_client.hget(RedisKey.user(g.user_id), 'current_org_id')})
        _body.update({"group_id": _suite.group_id})
        _id = Insert(Case, _body).insert_id(Case, "/case")
        return jsonify(error_code=RET.OK, error_msg="OK", data={"id": _id})

    @staticmethod
    @collect_sql_error
    def create_case_node_commit(body):
        _body = body.__dict__

        _suite = Suite.query.filter_by(name=_body.get("suite")).first()
        if not _suite:
            return jsonify(
                error_code=RET.PARMA_ERR,
                error_msg="The suite {} is not exist".format(
                    _body.get("suite")
                )
            )
        _body["suite_id"] = _suite.id
        _body.pop("suite")
        _body.update({"creator_id": g.user_id})
        _body.update({"org_id": redis_client.hget(RedisKey.user(g.user_id), 'current_org_id')})

        _id = Insert(Case, _body).insert_id(Case, "/case")

        _case_node_body = CaseNodeBodySchema(
            case_id=_id,
            group_id=_body.get("group_id"),
            parent_id=_body.get("parent_id"),
            title=_body.get("name"),
            type="case",
            creator_id=g.user_id,
            org_id=redis_client.hget(RedisKey.user(g.user_id), 'current_org_id'),
        )

        _result = CaseNodeHandler.create(_case_node_body)
        _result = _result.get_json()
        _case_node = CaseNode.query.filter_by(id=_result["data"]).first()
        source = list()
        cur = _case_node

        while cur:
            if not cur.parent.all():
                source.append(cur.id)
                break
            if len(cur.parent.all()) > 1:
                raise RuntimeError(
                    "case_node should not have parents beyond one")

            source.append(cur.id)
            cur = cur.parent[0]

        return jsonify(error_code=RET.OK, error_msg="OK")


class TemplateCasesHandler:
    @staticmethod
    @collect_sql_error
    def get_all(git_repo_id):
        _filter = GetAllByPermission(GitRepo).get_filter()
        _gitrepo = GitRepo.query.filter(*_filter).first()
        if not _gitrepo:
            raise RuntimeError("gitrepo does not exist/ has no right")
        suites = Suite.query.join(Case).filter(
            Suite.git_repo_id == git_repo_id
        ).all()
        data = [suite.relate_case_to_json() for suite in suites]
        return jsonify(
            error_code=RET.OK,
            error_msg="OK",
            data=data,
        )


class HandlerCaseReview(object):
    @staticmethod
    @collect_sql_error
    def create(body: AddCaseCommitSchema):
        """发起"""
        case = Case.query.get(body.case_detail_id)
        if not case:
            return jsonify(error_code=RET.NO_DATA_ERR, error_msg="case not exists")
        case.machine_type = body.machine_type
        case.machine_num = body.machine_num
        case.preset = body.preset
        case.description = body.case_description
        case.steps = body.steps
        case.expectation = body.expectation
        case.remark = body.remark
        case.add_update()
        return jsonify(error_code=RET.OK, error_msg="OK")


class ResourceItemHandler:
    _days_dict = {
        'week': 7,
        'halfMonth': 15,
        'month': 30,
    }

    def __init__(self, commit_type="week", case_node:CaseNode=None, **kwargs):
        if kwargs.get("_type") and kwargs.get("_type") not in ['group', 'org']:
            raise RuntimeError("type error, parameter type must be group or org")
        if (
            kwargs.get("_type") == 'group' and not kwargs.get("group_id")
        ) or (
            kwargs.get("_type") == 'org' and not kwargs.get("org_id")
        ):
            raise RuntimeError("type not match, group_id/org_id not exist")
        if kwargs.get("case_node_type") == 'baseline' and not kwargs.get("baseline_id"):
            raise RuntimeError("baseline id should be offered")

        self._type = kwargs.get("_type")
        self.org_id = kwargs.get("org_id")
        self.group_id = kwargs.get("group_id")
        self.baseline_id = kwargs.get("baseline_id")
        self.case_node_type = kwargs.get("case_node_type")
        self.commit_type = commit_type
        self.case_node = case_node

    @staticmethod
    def _transform(data):
        steps = {}
        for (key, value) in data:
            step = {key: value}
            steps.update(step)
        return steps

    def get_case(self):
        cases_filter = self._get_table_filter(Case)
        cases_filter.append(CaseNode.type == 'case')

        if self.case_node_type == 'baseline':
            cases_filter.append(CaseNode.baseline_id == self.baseline_id)
        else:
            cases_filter.append(CaseNode.baseline_id == sqlalchemy.null())
        
        all_count = Case.query.join(CaseNode).filter(
            *cases_filter
        ).count()
        auto_count = Case.query.join(CaseNode).filter(
            *cases_filter, 
            Case.automatic.is_(True)
        ).count()

        auto_ratio = '0%'

        if all_count != 0:
            auto_ratio = str(math.floor((auto_count / all_count) * 100)) + '%'

        return all_count, auto_ratio

    def get_suite(self):
        suites_filter = self._get_table_filter(Suite)
        suites_filter.append(CaseNode.type == 'suite')

        if self.case_node_type == 'baseline':
            suites_filter.append(CaseNode.baseline_id == self.baseline_id)
        else:
            suites_filter.append(CaseNode.baseline_id == sqlalchemy.null())
        
        all_count = Suite.query.join(CaseNode).filter(
            *suites_filter,
        ).count()

        return all_count

    def count_children_case(self, node):
        if not node.children:
            return 0
        res = 0
        for child in node.children:
            if child.type == "case":
                res += 1
            else:
                res += self.count_children_case(child)
            
        return res

    def _get_table_filter(self, table): 
        if self._type == 'group':
            table_filter = [
                table.permission_type == 'group',
                table.org_id == int(redis_client.hget(RedisKey.user(g.user_id), 'current_org_id')),
                table.group_id == self.group_id
            ]
        else:
            table_filter = [
                table.permission_type == 'org',
                table.org_id == self.org_id
            ]
        return table_filter

    def get_case_distribute(self):
        first_children =  list(filter(lambda child:(
            child.type == "directory" or child.type == "suite"
        ), self.case_node.children))
        result = list()

        for first_child in first_children:
            res_first_count = self.count_children_case(first_child)

            result.append({
                "name": first_child.title,
                "value": res_first_count,
            })
            
        return result

    def run(self):
        case_count, auto_ratio = self.get_case()
        suite_count = self.get_suite()
        return_data = {
            "case_count": case_count,
            "auto_ratio": auto_ratio,
            "suite_count": suite_count,
        }     

        return jsonify(
            error_code=RET.OK,
            error_msg='OK',
            data=return_data
        )


class CaseSetHandler:
    @staticmethod
    @collect_sql_error
    def get_caseset_details(casenode, case_node_type, query):
        if case_node_type == "directory":
            if casenode.permission_type == "group":
                resource = ResourceItemHandler(
                    _type='group',
                    group_id=casenode.group_id,
                    case_node_type="directory",
                    commit_type=query.commit_type,
                    case_node=casenode,
                )
            if casenode.permission_type == "org":
                resource = ResourceItemHandler(
                    _type='org',
                    org_id=casenode.org_id,
                    case_node_type="directory",
                    commit_type=query.commit_type,
                    case_node=casenode
                )
        
            case_count, auto_ratio = resource.get_case()
            suite_count = resource.get_suite()
            return_data = {
                "case_count": case_count,
                "suite_count": suite_count,
                "auto_ratio": auto_ratio,
            }

            type_distribute = resource.get_case_distribute()
            return_data["type_distribute"] = type_distribute
        
        if case_node_type == "baseline":
            if casenode.permission_type == "group":
                resource = ResourceItemHandler(
                    _type='group',
                    group_id=casenode.group_id,
                    case_node_type="baseline",
                    baseline_id=casenode.baseline_id,
                    case_node=casenode,
                )
            if casenode.permission_type == "org":
                resource = ResourceItemHandler(
                    _type='org',
                    org_id=casenode.org_id,
                    case_node_type="baseline",
                    baseline_id=casenode.baseline_id,
                    case_node=casenode,
                )
        
            case_count, auto_ratio = resource.get_case()
            suite_count = resource.get_suite()
            return_data = {
                "case_count": case_count,
                "suite_count": suite_count,
                "auto_ratio": auto_ratio,
            }

            type_distribute = resource.get_case_distribute()
            return_data["type_distribute"] = type_distribute
        
        return return_data


class SuiteDocumentHandler:
    @staticmethod
    @collect_sql_error
    def post(suite_id, body):
        _body = body.__dict__
        suites = list()
        suite = Suite.query.filter_by(id=suite_id).first()
        if not suite:
            return jsonify(
                error_code=RET.PARMA_ERR,
                error_msg="The suite is not exist"
            )
        suites.append(suite)
        _body.update({
            "creator_id": g.user_id,
            "org_id": redis_client.hget(RedisKey.user(g.user_id), 'current_org_id'),
            "group_id": suite.group_id,
            "suite_id": suite_id,
            "permission_type": suite.permission_type,
            }
        )
        _id = Insert(SuiteDocument, _body).insert_id(SuiteDocument, "/suite_document")
        return jsonify(error_code=RET.OK, error_msg="OK", data={"id": _id})


class OrphanSuitesHandler:
    def __init__(self, query) -> None:
        self.current_org_id = redis_client.hget(
            RedisKey.user(g.user_id),
            "current_org_id",
        )

        self.filter = [
            Suite.org_id == int(self.current_org_id),
            Suite.permission_type == "org",
            CaseNode.id == sqlalchemy.null(),
        ]

        self.query = query
        self._add_params()

    def _add_params(self):
        _filter_params = []
        if self.query.name:
            _filter_params.append(Suite.name.like(f"%{self.query.name}%"))
        if self.query.owner:
            _filter_params.append(Suite.owner == self.query.owner)
        if self.query.git_repo_url:
            _filter_params.append(Suite.git_repo.git_url == self.query.git_repo_url)
        if self.query.framework_name:
            _filter_params.append(Suite.framework.name == self.query.framework_name)
        self.filter += _filter_params

    def add_filters(self, query_filters: list):
        self.filter += query_filters

    @collect_sql_error
    def get_all(self):
        query_filter = Suite.query.outerjoin(CaseNode).filter(*self.filter).order_by(
            Suite.name, 
            Suite.create_time
        )

        def page_func(item):
            suite_dict = item.to_json()
            return suite_dict
        
        page_dict, e = PageUtil.get_page_dict(
            query_filter, 
            self.query.page_num, 
            self.query.page_size, 
            func=page_func
        )
        if e:
            return jsonify(error_code=RET.SERVER_ERR, error_msg=f'get orphan suites page error {e}')
        return jsonify(error_code=RET.OK, error_msg="OK", data=page_dict)


class CasefileExportUtil:
    COL_NAMES = [
        '测试套', '用例名', '测试级别', '测试类型', '用例描述', '节点数', 
        '预置条件', '操作步骤', '预期输出', '是否自动化', '备注'
    ]

    def __init__(self, filename: str, cases: List[dict]):
        self.filename = filename
        self.cases = cases

    @abc.abstractmethod
    def create_casefile(self):
        pass

    @abc.abstractmethod
    def add_row(self):
        pass

    def get_casefile(self):
        return self.casefile
    
    def rm_casefile(self):
        if os.path.isfile(self.casefile):
            os.remove(self.casefile)

    def inject_data(self):
        if not self.casefile:
            return
        for case_ in self.cases:
            self._inject_row(case_)

    def _inject_row(self, case: dict):
        _row_data = self._extract_row(case)
        self.add_row(_row_data)

    def _extract_row(self, case: dict):
        _row = []
        for col_name in self.COL_NAMES:
            translation_dict = current_app.config.get("OE_QA_TESTCASE_DICT")
            try:
                _v = case.get(translation_dict.get(col_name))
                if not _v:
                    _v = ''
                else:
                    _v = str(_v)
                _row.append(_v)
            except KeyError:
                _row.append('')
        return _row


class ExcelExportUtil(CasefileExportUtil):
    def __init__(self, filename: str, cases: List[dict]):
        self.next_row = 1
        super().__init__(filename, cases)
        self.casefile = os.path.join(current_app.config.get("TMP_FILE_SAVE_PATH"), "export", f"{self.filename}.xlsx")

    def create_casefile(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = self.filename
        for i, col_name in enumerate(self.COL_NAMES):
            sheet.cell(self.next_row, i+1).value = col_name
        wb.save(self.casefile)
        self.next_row += 1
    
    def add_row(self, row_data: list):
        wb = openpyxl.load_workbook(self.casefile)
        sheet = wb[self.filename]
        for i, value in enumerate(row_data):
            sheet.cell(self.next_row, i+1).value = value
            sheet.cell(self.next_row, i+1).alignment = Alignment(
                wrapText=True,
                vertical='top',
                horizontal='left',
            )
        
        wb.save(self.casefile)
        self.next_row += 1


class MdExportUtil(CasefileExportUtil):
    def __init__(self, filename: str, cases: List[dict]):
        super().__init__(filename, cases)
        self.casefile = os.path.join(current_app.config.get("TMP_FILE_SAVE_PATH"), "export", f"{self.filename}.md")

    def create_casefile(self):
        with open(self.casefile,  "a") as f:
            f.write(f"|{'|'.join(self.COL_NAMES)}|\n")
            split_row = "|---"*len(self.COL_NAMES) + "|\n"
            f.write(split_row)
    
    def add_row(self, row_data: list):
        with open(self.casefile, "a+") as f:
            _row = '|'
            for cell in row_data:
                cell = cell.replace('\n', '<br />')
                _row = _row + str(cell) + '|'
            _row += '\n'
            f.write(_row)


class TestResultEventHandler:
    @staticmethod
    def post_test_result(body):
        # 校验
        if not body.result:
            return jsonify(
                error_code=RET.PARMA_ERR,
                error_msg="Organization is not existed!",
            )
        org = Organization.query.filter_by(name=body.org).first()
        if not org:
            return jsonify(
                error_code=RET.PARMA_ERR,
                error_msg="Organization is not existed!",
            )
        suite = Suite.query.filter_by(name=body.testsuite).first()
        if not suite:
            return jsonify(
                error_code=RET.PARMA_ERR,
                error_msg="suite is not existed!",
            )

        case = Case.query.filter_by(name=body.testcase).first()
        if not case:
            return jsonify(
                error_code=RET.PARMA_ERR,
                error_msg="case is not existed!",
            )

        baseline_node = CaseNode.query.filter_by(org_id=org.id, type="baseline", is_root=True, in_set=False,
                                                 title=body.baseline, permission_type="org").first()
        if not baseline_node:
            return jsonify(
                error_code=RET.PARMA_ERR,
                error_msg="baseline is not existed!",
            )
        case_node = CaseNode.query.filter_by(org_id=org.id, type="case", is_root=False, in_set=False, case_id=case.id,
                                             baseline_id=baseline_node.baseline_id, permission_type="org").first()
        if not case_node:
            return jsonify(
                error_code=RET.PARMA_ERR,
                error_msg="case node is not existed!",
            )
        result_data = {
            "case_id": case.id,
            "baseline_id": baseline_node.baseline_id,
            "result": body.result,
            "details": body.details,
            "log_url": body.log_url,
            "fail_type": body.fail_type,
            "running_time": body.running_time,
        }
        case_result = CaseResult.query.filter_by(baseline_id=baseline_node.baseline_id, case_id=case.id).first()
        if case_result:
            result_data.update({"id": case_result.id})
            Edit(CaseResult, result_data).single(CaseResult, "/case_result")
        else:
            Insert(CaseResult, result_data).single()

        return jsonify(
            error_code=RET.OK,
            error_msg="OK",
        )

    @staticmethod
    def get_test_result(query):
        baseline = Baseline.query.filter_by(milestone_id=query.milestone_id).first()
        data = {}
        if baseline:
            case_result = CaseResult.query.filter_by(case_id=query.case_id, baseline_id=baseline.id).first()
            if case_result:
                data = case_result.to_json()

        return jsonify(
            data=data,
            error_code=RET.OK,
            error_msg="OK",
        )

    @staticmethod
    def test_result(body):
        if not body.result:
            return jsonify(
                error_code=RET.PARMA_ERR,
                error_msg="Organization is not existed!",
            )
        org = Organization.query.filter_by(id=body.org_id).first()
        if not org:
            return jsonify(
                error_code=RET.PARMA_ERR,
                error_msg="Organization is not existed!",
            )

        case = Case.query.filter_by(id=body.case_id).first()
        if not case:
            return jsonify(
                error_code=RET.PARMA_ERR,
                error_msg="case is not existed!",
            )

        baseline = Baseline.query.filter_by(milestone_id=body.milestone_id).first()
        if not baseline:
            return jsonify(
                error_code=RET.PARMA_ERR,
                error_msg="baseline is not existed!",
            )

        case_node = CaseNode.query.filter_by(org_id=org.id, type="case", is_root=False, in_set=False,
                                             baseline_id=baseline.id, permission_type="org", case_id=case.id).first()
        if not case_node:
            return jsonify(
                error_code=RET.PARMA_ERR,
                error_msg="case node is not existed!",
            )
        result_data = {
            "case_id": case.id,
            "baseline_id": baseline.id,
            "result": body.result,
            "details": body.details,
            "log_url": body.log_url,
            "fail_type": body.fail_type,
            "running_time": body.running_time,
        }
        case_result = CaseResult.query.filter_by(baseline_id=baseline.id, case_id=case.id).first()
        if case_result:
            result_data.update({"id": case_result.id})
            Edit(CaseResult, result_data).single(CaseResult, "/case_result")
        else:
            Insert(CaseResult, result_data).single()
        return jsonify(
            error_code=RET.OK,
            error_msg="OK",
        )

# Copyright (c) [2022] Huawei Technologies Co.,Ltd.ALL rights reserved.
# This program is licensed under Mulan PSL v2.
# You can use it according to the terms and conditions of the Mulan PSL v2.
#          http://license.coscl.org.cn/MulanPSL2
# THIS PROGRAM IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
# EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
# MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
# See the Mulan PSL v2 for more details.
####################################
# @Author  :
# @email   :
# @Date    :
# @License : Mulan PSL v2
#####################################

import json
from openpyxl import Workbook

from lib.common import RestApi


def export_case_to_excel(filepath: str, auth: str=None) -> None:
    """
        导出case详细信息到excel文档(.xlsx)
    """
    wb = Workbook()
    ws = wb.active
    api = RestApi("/api/v1/case", auth=auth)
    resp = api.get(verify=False)
    resp_dict = json.loads(resp.text)
    case_list = resp_dict.get("data")
    # 表头
    ws.cell(1, 1, "测试套")
    ws.cell(1, 2, "用例名")
    ws.cell(1, 3, "测试级别")
    ws.cell(1, 4, "测试类型")
    ws.cell(1, 5, "用例描述")
    ws.cell(1, 6, "节点数")
    ws.cell(1, 7, "预置条件")
    ws.cell(1, 8, "操作步骤")
    ws.cell(1, 9, "预期输出")
    ws.cell(1, 10, "是否自动化")
    ws.cell(1, 11, "备注")
    # 内容
    row = 2
    for _case in case_list:
        ws.cell(row, 1, _case.get("suite"))
        ws.cell(row, 2, _case.get("name"))
        ws.cell(row, 3, _case.get("test_level"))
        ws.cell(row, 4, _case.get("test_type"))
        ws.cell(row, 5, _case.get("description"))

        ws.cell(row, 7, _case.get("preset"))
        ws.cell(row, 8, _case.get("steps"))
        ws.cell(row, 9, _case.get("expection"))
        ws.cell(row, 10, _case.get("automatic"))
        ws.cell(row, 11, _case.get("remark"))

        # 计算节点数
        case_id = _case.get("id")
        node_num = 0
        api = RestApi("/api/v1/case-node")
        resp = api.get(verify=False)
        resp_dict = json.loads(resp.text)
        case_node_list = resp_dict.get("data")
        for _case_node in  case_node_list:
            if _case_node.get("case_id") == case_id:
                node_num = node_num + 1
        ws.cell(row, 6, node_num)

        row = row + 1

    wb.save(filepath)

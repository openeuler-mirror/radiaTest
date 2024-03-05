# Copyright (c) [2022] Huawei Technologies Co.,Ltd.ALL rights reserved.
# This program is licensed under Mulan PSL v2.
# You can use it according to the terms and conditions of the Mulan PSL v2.
#          http://license.coscl.org.cn/MulanPSL2
# THIS PROGRAM IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
# EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
# MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
# See the Mulan PSL v2 for more details.
####################################
# @Author  : disnight
# @email   : fjc837005411@outlook.com
# @Date    : 2022/08/21
# @License : Mulan PSL v2
#####################################

from glob import escape
import io
import html
from bcrypt import re

from flask import current_app
import openpyxl
import markdown
from lxml import etree


class MdUtil:

    DEFAULT_MD_EXT = [
        "markdown.extensions.extra",
        "markdown.extensions.codehilite"
    ]

    @staticmethod
    def get_md_tables2html(md_content):
        """parse md table 2 lxml table list

        Args:
            md_content (str): content of markdown file
        
        return:
            list: [ lxml_table_1, lxml_table_2]
        """
        try :
            html_content = markdown.markdown(md_content, extensions=MdUtil.DEFAULT_MD_EXT)
        except UnicodeDecodeError as ude:
            current_app.logger.error(f"unsupported unicode in markdown file :{str(ude)}")
            html_content = ""
        except ValueError as ve:
            current_app.logger.error(f"unsupported grammar in markdown file :{str(ve)}")
            html_content = ""
        
        if not html_content:
            return []

        html_content = html.escape(html_content)
        html_etree = etree.HTML(html_content, parser=etree.HTMLParser(encoding="utf-8"))
        table_list = html_etree.xpath("//table")
        return table_list

    @staticmethod
    def md2html(md_content, file_path):
        """convert md text to html file

        Args:
            md_content (str): content of markdown file
            file_path (str): html file path

        """
        try :
            html_content = markdown.markdown(md_content, output_format="html", extensions=MdUtil.DEFAULT_MD_EXT)
        except UnicodeDecodeError as ude:
            current_app.logger.error(f"unsupported unicode in markdown file :{str(ude)}")
            html_content = ""
        except ValueError as ve:
            current_app.logger.error(f"unsupported grammar in markdown file :{str(ve)}")
            html_content = ""
        
        if not html_content:
            return
        html_content = html.escape(html_content)

        fhtml = io.open(file_path, "w", encoding="utf-8")
        fhtml.write(html_content)
        fhtml.close()

    @staticmethod
    def get_md_tables2list(md_content, resolver):
        """parse md table 2 python list

        Args:
            md_content (str): content of markdown file
        
        return:
            list: [[md table 1],[md table 2]]
        """
        table_list = MdUtil.get_md_tables2html(md_content)

        tables_content_list = []
        for table in table_list:
            tables_content_list.append(
                resolver(table).parse_table()
            )
        return tables_content_list

    @staticmethod
    def df2md(df, md_path):
        """convert pandas dataframe to md text

        Args:
            df (List(Dict[colname, value])): dataframe object loaded
            md_path (str): markdown file path
        
        Return:
            filepath (str): the convert file save path 

        """
        md_content = ""
        title = "|"
        split_line = "|"

        for col_name in df[0].keys():
            title = f"{title} {col_name} |"
            split_line = f"{split_line} -- |"
        
        md_content += title + "\n"
        md_content += split_line + "\n"

        for row in df:
            row_content = "|"
            for col in row.values():
                cell_content = str(col).replace("\n", "<br/>").replace("nan", "")
                row_content += cell_content + "|"
            
            md_content += row_content + "\n"
        
        with open(md_path, "w") as file:
            file.write(md_content)
        
        return md_path

    @staticmethod
    def md2wb(md_content, wb_path, sheet_name):
        """convert md text to excel workbook

        Args:
            md_content (str): source markdown content which only has single table
            wb_path (str): target workbook path (filepath should be valid)
            sheet_name (str): name of the target sheet of saving workbook
        
        Return:
            filepath (str): the convert file save path 

        """

        escape_dict = {
            "<br>": "\n",
            "<br/>": "\n",
            "\<": "<",
            "\\$": "\$",
            "\|": "|",
        }

        md_lines = md_content.split("\n")
        # 定位表头
        header = 0
        split_pattern = r"^(\|\s*-+\s*)+\|$"
        for i in range(1, len(md_lines) - 1):
            if re.match(split_pattern, md_lines[i]):
                if len(md_lines[i - 1].strip("|").split("|")) == len(md_lines[i].strip("|").split("|")):
                    header = i - 1
        
        title = md_lines[header].strip("|").split("|")
        cols_num = len(title)

        wb = openpyxl.load_workbook(wb_path)
        ws = wb[sheet_name]
        # 表头赋值
        for col in range(cols_num):
            ws.cell(1, col + 1).value = title[col]
        
        # 搜索有效表行
        row_pattern = r"^" + "".join([
            '(?:\|((?!\s*(?:grep|awk|sed|tee|sort|uniq|tail|more|less)).*?[^\\\]{1,50})?)' for _ in range(cols_num)
        ]) + "\|$"
        row_index = 2

        def escape_content(cell_content):
            if not cell_content:
                cell_content = ''
            escape_content = cell_content.strip()
            for key, value in escape_dict.items():
                escape_content = escape_content.replace(key, value)
            return escape_content

        for i in range(header + 2, len(md_lines)):
            result = re.match(row_pattern, md_lines[i])
            if result:
                cells = result.groups()
                # md转行替换，格式转换
                body = list(map(escape_content, cells))
                # 表行赋值
                if len(body) == cols_num:
                    for col in range(cols_num):
                        ws.cell(row_index, col + 1).value = body[col]
                    row_index += 1

        wb.save(wb_path)

        return wb_path
